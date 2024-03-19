# -*- coding: utf-8 -*-
from datetime import datetime
from odoo import api, fields, models, _
import io
import base64
import openpyxl
from translate import Translator
import subprocess

class ZafcoPrisma(models.Model):
    _name = "zf.prisma"
    _description = "prisma excel report"
    _rec_name = 'name'

    name = fields.Char(string="Name", required=True)
    zf_translate_line_ids = fields.One2many('zf.translate.lines', 'translate_id')
    prisma_file_xlsx = fields.Binary(string='Excel File')
    prisma_translate_file_xlsx = fields.Binary(string='Translated Excel File')

    def action_button_translate(self):
        translated_excel_data = io.BytesIO()

        output_wb = openpyxl.Workbook()
        output_sheet = output_wb.active
        output_sheet.title = 'Sheet1'

        excel_data = io.BytesIO(base64.b64decode(self.prisma_file_xlsx))
        wb = openpyxl.load_workbook(excel_data)

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        if isinstance(cell.value, str):
                            existing_record = self.env['zf.prisma.data'].search([('source_data', '=', cell.value)])
                            if existing_record:
                                translated_text = existing_record.destination_data
                            else:
                                translator = Translator(to_lang="en", from_lang="ru")
                                translated_text = translator.translate(cell.value)
                                self.env['zf.prisma.data'].create({
                                    'source_data': cell.value,
                                    'destination_data': translated_text
                                })
                            output_sheet[cell.coordinate] = translated_text
                        elif isinstance(cell.value, (int, float)):
                            output_sheet[cell.coordinate] = cell.value
                    else:
                        output_sheet[cell.coordinate] = ""

            output_wb.save(translated_excel_data)

        translated_excel_data.seek(0)

        excel_binary_data = base64.b64encode(translated_excel_data.getvalue())
        self.write({'prisma_translate_file_xlsx': excel_binary_data})
        excel_data = io.BytesIO(base64.b64decode(self.prisma_translate_file_xlsx))
        wb = openpyxl.load_workbook(excel_data, data_only=True)

        translate_lines = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            max_row = sheet.max_row
            for row_index, row in enumerate(sheet.iter_rows(min_row=9, values_only=True), start=9):
                if row_index == max_row:
                    continue
                date_str = row[0]
                date_obj = datetime.strptime(date_str, '%d.%m.%Y').date()
                line_data = {
                    'translate_id': self.id,
                    'date': date_obj,
                    'document_name': row[1],
                    'analytic_debit': row[3],
                    'analytic_credit': row[4],
                    'debit_code': row[5],
                    'debit_amount': float(row[6]) if isinstance(row[6], (int, float)) else 0.0,
                    'credit_code': row[8],
                    'credit_amount': float(row[9]) if isinstance(row[9], (int, float)) else 0.0
                }
                translate_lines.append((0, 0, line_data))
        self.write({'zf_translate_line_ids': translate_lines})

class ZafcoTranslateLines(models.Model):
    _name = "zf.translate.lines"
    _description = "Zafco Translate Lines"

    translate_id = fields.Many2one('zf.prisma', string="Roaster Id")
    date = fields.Date(string="Date")
    document_name = fields.Char(string="Name")
    analytic_debit = fields.Char(string="Analytic Debit")
    analytic_credit = fields.Char(string="Analytic Credit")
    debit_code = fields.Integer(string="Debit Code")
    debit_amount = fields.Float(string="Debit Amount")
    credit_code = fields.Integer(string="Credit Code")
    credit_amount = fields.Float(string="Credit Amount")
